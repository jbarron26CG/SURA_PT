import streamlit as st
import pandas as pd
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from googleapiclient.errors import HttpError
from gspread.exceptions import APIError
import io
import gspread
import bcrypt
#import datetime
from datetime import datetime
import re
from zoneinfo import ZoneInfo
import yagmail
import time
from supabase import create_client
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import altair as alt
import numpy as np

supabase = create_client(
    st.secrets["SUPABASE_URL"],
    st.secrets["SUPABASE_KEY"]
)

# =======================================================
#             CONFIGURAR CREDENCIALES
# =======================================================
scope = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

# 🔥 CREDENCIALES DESDE STREAMLIT SECRETS
creds = Credentials.from_service_account_info(
    st.secrets["gcp_service_account"],
    scopes=scope
)

client = gspread.authorize(creds)
drive_service = build("drive", "v3", credentials=creds)


# =======================================================
#     FUNCIONES PARA GOOGLE DRIVE DENTRO DE SHARED DRIVE
# =======================================================

#id de pruebas
#SHARED_DRIVE_ID = "0AMMvXls255mTUk9PVA"   # ← tu unidad compartida

#ID unidad real
SHARED_DRIVE_ID = "0AMe71RDJTYcGUk9PVA"

def obtener_o_crear_carpeta(nombre_carpeta, drive_service):
    """Busca o crea carpeta en la unidad compartida."""

    query = (
        f"name = '{nombre_carpeta}' "
        f"and mimeType = 'application/vnd.google-apps.folder' "
        f"and '{SHARED_DRIVE_ID}' in parents "
        f"and trashed = false"
    )

    resultado = drive_service.files().list(
        q=query,
        spaces='drive',
        corpora='drive',
        driveId=SHARED_DRIVE_ID,
        includeItemsFromAllDrives=True,
        supportsAllDrives=True,
        fields="files(id, name)"
    ).execute()

    folders = resultado.get("files", [])

    if folders:
        return folders[0]["id"]

    metadata = {
        "name": nombre_carpeta,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [SHARED_DRIVE_ID]
    }

    nueva = drive_service.files().create(
        body=metadata,
        fields="id",
        supportsAllDrives=True
    ).execute()

    return nueva["id"]

def obtener_carpeta(nombre_carpeta, drive_service):
    """Busca una carpeta en la unidad compartida.
    Devuelve el ID si existe, o None si no existe.
    """

    query = (
        f"name = '{nombre_carpeta}' "
        f"and mimeType = 'application/vnd.google-apps.folder' "
        f"and '{SHARED_DRIVE_ID}' in parents "
        f"and trashed = false"
    )

    resultado = drive_service.files().list(
        q=query,
        spaces="drive",
        corpora="drive",
        driveId=SHARED_DRIVE_ID,
        includeItemsFromAllDrives=True,
        supportsAllDrives=True,
        fields="files(id, name)"
    ).execute()

    folders = resultado.get("files", [])

    if folders:
        return folders[0]["id"]

    return None


def subir_archivo_drive(nombre_archivo, contenido, mime_type, folder_id, drive_service):
    """Sube un archivo dentro de una carpeta en Shared Drive."""

    file_metadata = {
        "name": nombre_archivo,
        "parents": [folder_id]
    }

    media = MediaIoBaseUpload(io.BytesIO(contenido), mimetype=mime_type)

    archivo = drive_service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id",
        supportsAllDrives=True
    ).execute()

    return archivo["id"]

# =======================================================
#                       LOGIN
# =======================================================
def login():

    st.title("Inicio de Sesión")

    user = st.text_input("USUARIO:")
    password = st.text_input("CONTRASEÑA:", type="password")

    if user or password:
        ingreso = st.button("Ingresar",use_container_width=True)
        if ingreso:
            if not user or not password:
                st.warning("Ingresa usurio y contraseña")
                return
            response = (
                supabase
                .table("Login")
                .select("USUARIO","PASSWORD","ROL","LIQUIDADOR")
                .eq("USUARIO",user)
                .limit(1)
                .execute()
            )

            if not response.data:
                st.error("Usuario o contraseña incorrectos")
                return
            
            registro = response.data[0]

            flag_psw = bcrypt.checkpw(password.encode("utf-8"), registro["PASSWORD"].encode("utf-8"))
            time.sleep(1)
            if not flag_psw:
                st.error("Usuario o contraseña incorrectos.")
                return
            st.session_state["USUARIO"] = registro["USUARIO"]
            st.session_state["ROL"] = registro["ROL"]
            st.session_state["LIQUIDADOR"] = registro["LIQUIDADOR"]
            st.session_state["auth"] = True

            st.success(f"Acceso exitoso")
            st.rerun()
    else:
        ingreso = st.button("Ingresar",use_container_width=True,disabled=True)


def reset_form_registro():
    """Reinicia los valores del formulario de registro de siniestro en st.session_state."""
    
    # Define los valores por defecto exactos que usas en el bloque de inicialización
    default_values = {
        "siniestro_num": "",
        "siniestro_correl": "",
        "siniestro_fecha": datetime.today().date(),
        "siniestro_lugar": "",
        "siniestro_medio": "Call center",
        "Cobertura":"",
        "aseg_nombre": "",
        "aseg_rut": "",
        "aseg_tipo": "",
        "aseg_tel": "",
        "aseg_correo": "",
        "aseg_direccion": "",
        "prop_nombre": "",
        "prop_rut": "",
        "prop_tipo": "",
        "prop_tel": "",
        "prop_correo": "",
        "prop_direccion": "",
        "veh_marca": "",
        "veh_submarca": "",
        "veh_version": "",
        "veh_anio": "",
        "veh_serie": "",
        "veh_motor": "",
        "veh_patente": "" 
    }
    
    # Itera sobre los valores por defecto y actualiza st.session_state
    # Solo actualiza si la clave está actualmente en st.session_state (para seguridad)
    for key, default_value in default_values.items():
        if key in st.session_state:
            st.session_state[key] = default_value
    if "veh_archivos" in st.session_state:
        del st.session_state["veh_archivos"]

def limpiar_y_recargar():
    reset_form_registro()


def panel_subir_documentos():
    st.subheader("⬆️ Subir archivos a drive")
    siniestro_id = st.text_input("ESCRIBE NÚMERO DE SINIESTRO")

    if not siniestro_id:
        st.info("Ingresa un número para buscar un siniestro.")
        return
    nombre_carpeta = f"SINIESTRO_{siniestro_id}"
    carpeta_id = obtener_carpeta(nombre_carpeta, drive_service)
    if carpeta_id is None:
        st.warning("⚠️ No existe carpeta para este siniestro, verifica información ingresada")
        return
    with st.form("form_seguimiento", clear_on_submit=True):
        uploaded_files = st.file_uploader(
            "Selecciona los archivos",
            accept_multiple_files=True
        )

        enviado = st.form_submit_button("Cargar archivos",icon="💾",use_container_width=True)
    if enviado:
        if uploaded_files:
            for archivo in uploaded_files:
                subir_archivo_drive(
                    archivo.name,
                    archivo.read(),
                    archivo.type,
                    carpeta_id,
                    drive_service
                )
        st.toast("Guardando cambios...", icon="⏳",duration=1)
        time.sleep(1)
        st.toast("Archivos cargados correctamente", icon="✅")
        st.success("Archivos cargados correctamente", icon="✅")
    
def panel_seguimiento(siniestro_id):
    st.subheader("📌 Agregar Estatus (Seguimiento)")
    response = (
        supabase
        .table("BitacoraOperaciones")
        .select("*")
        .eq("NUM_SINIESTRO",siniestro_id)
        .execute()
    )
    with st.form("form_seguimiento", clear_on_submit=True):

        nuevo_estatus = st.selectbox(
            "ESTATUS",
            [
                "Seleccionar estatus",
                "ASIGNADO","CLIENTE CONTACTADO","CARGA DOCUMENTAL RECIBIDA",
                "DESVIADO A FRAUDES","DOCUMENTACIÓN COMPLETA",
                "EN ESPERA DE PRIMAS, PÓLIZA Y/O SALDO INSOLUTO",
                "PROPUESTA ECONÓMICA ENVIADA","PROPUESTA ECONÓMICA ACEPTADA",
                "DERIVADO A CERO KM","DERIVADO A REPOSICIÓN",
                "EN ESPERA DE PRIMERA FIRMA",
                "EN ESPERA DE SEGUNDA FIRMA (ROBO)",
                "EN ESPERA DE LEGALIZACIÓN","DOCUMENTACIÓN LEGALIZADA",
                "SOLICITUD DE PAGO GENERADA","PAGO LIBERADO",
                "CIERRE POR DESISTIMIENTO","CIERRE POR RECHAZO",
                "DERIVADO A PARCIALES"
            ]
        )

        comentario = st.text_area("COMENTARIOS", height=120)

        uploaded_files = st.file_uploader(
            "Selecciona los archivos",
            accept_multiple_files=True
        )

        enviado = st.form_submit_button("Agregar estatus",icon="💾",use_container_width=True)

    if enviado:

        if nuevo_estatus == "Seleccionar estatus":
            st.warning("Debes seleccionar un estatus.")
            return
        ref = response.data[-1]
        ahora = datetime.now(ZoneInfo("America/Mexico_City"))

        ref["FECHA_ESTATUS_BITACORA"] = ahora.strftime("%Y-%m-%d %H:%M:%S")
        ref["ESTATUS"] = nuevo_estatus
        ref["COMENTARIO"] = comentario
        ref["CORREO_LIQUIDADOR"] = st.session_state["USUARIO"]

        supabase.table("BitacoraOperaciones").insert(ref).execute()

        if uploaded_files:
            nombre_carpeta = f"SINIESTRO_{siniestro_id}"
            carpeta_id = obtener_o_crear_carpeta(nombre_carpeta, drive_service)

            for archivo in uploaded_files:
                subir_archivo_drive(
                    archivo.name,
                    archivo.read(),
                    archivo.type,
                    carpeta_id,
                    drive_service
                )

        st.session_state["last_load_time"] = 0
        st.toast("Guardando cambios...", icon="⏳",duration=1)
        time.sleep(1)
        st.toast("Estatus agregado correctamente", icon="✅",duration=1)
        st.success("Estatus agregado correctamente", icon="✅")
        time.sleep(1)
        st.rerun()


def panel_modificar_datos(siniestro_id):

    st.subheader("✏️ Modificar Datos")

    # Usamos la primera fila como referencia
    response = (
        supabase
        .table("BitacoraOperaciones")
        .select("*")
        .eq("NUM_SINIESTRO",siniestro_id)
        .execute()
    )
    ref = response.data[-1]
    fecha_creacion = ref["FECHA_CREACION"]
    # Campos a editar
    with st.expander("DATOS DEL SINIESTRO", expanded=False):
        num_siniestro = siniestro_id
        correlativo = st.text_input("Correlativo", ref["CORRELATIVO"])
        fecha_siniestro = st.date_input("Fecha del siniestro", pd.to_datetime(ref["FECHA_SINIESTRO"]))
        lugar = st.text_input("Lugar del siniestro", ref["LUGAR_SINIESTRO"])
        medio = st.selectbox("Medio de asignación", ["Call center", "PP", "ALMA"], index=["Call center","PP","ALMA"].index(ref.get("MEDIO") or ""))
        Cobertura = st.selectbox("Cobertura", ["","Robo", "Daño material"], index=["","Robo", "Daño material"].index(ref.get("COBERTURA") or ""))

    # Datos asegurado
    with st.expander("DATOS DEL ASEGURADO", expanded=False):
        asegurado_nombre = st.text_input("Nombre asegurado", ref["NOMBRE_ASEGURADO"])
        asegurado_rut = st.text_input("RUT asegurado", ref["RUT_ASEGURADO"])
        asegurado_tipo = st.selectbox("Tipo persona asegurado", ["","Jurídica", "Natural"], index=["","Jurídica", "Natural"].index(ref.get("TIPO_DE_PERSONA_ASEGURADO") or ""))
        asegurado_tel = st.text_input("Teléfono asegurado", ref["TEL_ASEGURADO"])
        asegurado_correo = st.text_input("Correo asegurado", ref["CORREO_ASEGURADO"])
        asegurado_dir = st.text_input("Dirección asegurado", ref["DIRECCION_ASEGURADO"])

    # Datos propietario
    with st.expander("DATOS DEL PROPIETARIO", expanded=False):
        propietario_nombre = st.text_input("Nombre propietario", ref["NOMBRE_PROPIETARIO"])
        propietario_rut = st.text_input("RUT propietario", ref["RUT_PROPIETARIO"])
        propietario_tipo = st.selectbox("Tipo persona propietario", ["","Jurídica", "Natural"], index=["","Jurídica", "Natural"].index(ref.get("TIPO_DE_PERSONA_PROPIETARIO") or ""))
        propietario_tel = st.text_input("Tel. propietario", ref["TEL_PROPIETARIO"])
        propietario_correo = st.text_input("Correo propietario", ref["CORREO_PROPIETARIO"])
        propietario_dir = st.text_input("Dirección propietario", ref["DIRECCION_PROPIETARIO"])

    # Datos vehículo
    with st.expander("DATOS DEL VEHÍCULO", expanded=False):
        marca = st.text_input("Marca", ref["MARCA"])
        submarca = st.text_input("Submarca", ref["SUBMARCA"])
        version = st.text_input("Versión", ref["VERSION"])
        anio = st.text_input("Año/Modelo", ref["MODELO"])
        serie = st.text_input("Número de serie", ref["NO_SERIE"])
        motor = st.text_input("Motor", ref["MOTOR"])
        patente = st.text_input("Patente", ref["PATENTE"])

    if st.button("💾 Guardar cambios", use_container_width=True):
        supabase.table("BitacoraOperaciones").update({
            "NUM_SINIESTRO": num_siniestro,
            "FECHA_CREACION": fecha_creacion,
            "CORRELATIVO": correlativo,
            "FECHA_SINIESTRO": fecha_siniestro.strftime("%Y-%m-%d"),
            "LUGAR_SINIESTRO": lugar,
            "MEDIO": medio,
            "COBERTURA": Cobertura,

            "NOMBRE_ASEGURADO":asegurado_nombre,
            "RUT_ASEGURADO": asegurado_rut,
            "TIPO_DE_PERSONA_ASEGURADO": asegurado_tipo,
            "TEL_ASEGURADO": asegurado_tel,
            "CORREO_ASEGURADO": asegurado_correo,
            "DIRECCION_ASEGURADO": asegurado_dir,

            "NOMBRE_PROPIETARIO": propietario_nombre,
            "RUT_PROPIETARIO": propietario_rut,
            "TIPO_DE_PERSONA_PROPIETARIO": propietario_tipo,
            "TEL_PROPIETARIO": propietario_tel,
            "CORREO_PROPIETARIO": propietario_correo,
            "DIRECCION_PROPIETARIO": propietario_dir,

            "MARCA": marca,
            "SUBMARCA": submarca,
            "VERSION": version,
            "MODELO": anio,
            "NO_SERIE": serie,
            "MOTOR": motor,
            "PATENTE": patente
        }).eq("NUM_SINIESTRO", siniestro_id).execute()

        st.session_state["last_load_time"] = 0
        st.toast("Guardando cambios...", icon="⏳",duration=1)
        time.sleep(1)
        st.toast("Datos actualizados correctamente", icon="✅",duration=1)
        st.success("Datos actualizados correctamente", icon="✅")
        time.sleep(1)
        st.rerun()

def vista_modificar_siniestro():

    st.subheader("🔍 Buscar siniestro para actualizar")

    busqueda = st.text_input("ESCRIBE NÚMERO DE SINIESTRO")

    if not busqueda:
        st.info("Ingresa un número para buscar un siniestro.")
        return
    try:
        response = (
            supabase
            .table("BitacoraOperaciones")
            .select("*")
            .eq("NUM_SINIESTRO", busqueda)
            .execute()
        )
    except Exception as e:
        st.error("Error al consultar el siniestro.")
        st.write(e)
        st.stop()

    if not response.data:
        st.error("Siniestro no encontrado.",icon="❌")
        return
    
    seleccionado = response.data[0].get("NUM_SINIESTRO")

    if not seleccionado:
        return

    st.session_state["siniestro_actual"] = seleccionado

    st.success(f"Siniestro encontrado", icon="✅")

    tab1, tab2 = st.tabs(["✏️ MODIFICAR DATOS", "📌 SEGUIMIENTO"])

    with tab1:
        panel_modificar_datos(seleccionado)

    with tab2:
        panel_seguimiento(seleccionado)

    if st.button("Volver al inicio", icon="⬅️", use_container_width=True):
        st.session_state.vista = None
        st.rerun()




def registro_siniestro():
    st.header("Registro de nuevo siniestro")

    with st.form("form_siniestro",clear_on_submit=True):

        tabs = st.tabs([
            "📁 Datos del siniestro",
            "👤 Datos del asegurado",
            "👤 Datos del propietario",
            "🚗 Datos del vehículo"
        ])

        # ---------------------- DATOS SINIESTRO -------------------------
        with tabs[0]:
            st.subheader("Datos del siniestro")

            Siniestro = st.text_input("Número de siniestro", key="siniestro_num")
            Correlativo = st.text_input("Correlativo", key="siniestro_correl")
            FechaSiniestro = st.date_input("Fecha del siniestro", key="siniestro_fecha")
            Lugar = st.text_input("Lugar del siniestro", key="siniestro_lugar")
            Medio = st.selectbox(
                "Medio de asignación",
                ["Call center", "PP", "ALMA"],
                key="siniestro_medio"
            )
            Cobertura = st.selectbox("Cobertura", ["","Robo", "Daño material"],key="Cobertura")

        # ---------------------- DATOS ASEGURADO -------------------------
        with tabs[1]:
            st.subheader("Datos del asegurado")

            Asegurado_Nombre = st.text_input("Nombre", key="aseg_nombre")
            Asegurado_Rut = st.text_input("RUT", key="aseg_rut")
            Asegurado_Tipo = st.selectbox(
                "Tipo de persona",
                ["","Natural", "Jurídica"],
                key="aseg_tipo"
            )
            Asegurado_Telefono = st.text_input("Teléfono", key="aseg_tel")
            Asegurado_Correo = st.text_input("Correo electrónico", key="aseg_correo")
            Asegurado_Direccion = st.text_input("Dirección", key="aseg_direccion")

        # ---------------------- DATOS PROPIETARIO -------------------------
        with tabs[2]:
            st.subheader("Datos del propietario")

            Propietario_Nombre = st.text_input("Nombre", key="prop_nombre")
            Propietario_Rut = st.text_input("RUT", key="prop_rut")
            Propietario_Tipo = st.selectbox(
                "Tipo de persona",
                ["","Natural", "Jurídica"],
                key="prop_tipo"
            )
            Propietario_Telefono = st.text_input("Teléfono", key="prop_tel")
            Propietario_Correo = st.text_input("Correo electrónico", key="prop_correo")
            Propietario_Direccion = st.text_input("Dirección", key="prop_direccion")

        # ---------------------- DATOS VEHÍCULO -------------------------
        with tabs[3]:
            st.subheader("Datos del vehículo")

            Marca = st.text_input("Marca", key="veh_marca")
            Submarca = st.text_input("Submarca", key="veh_submarca")
            Version = st.text_input("Versión", key="veh_version")
            AñoModelo = st.text_input("Año/Modelo", key="veh_anio")
            Serie = st.text_input("Número de serie", key="veh_serie")
            Motor = st.text_input("Motor", key="veh_motor")
            Patente = st.text_input("Patente", key="veh_patente")

            archivos = st.file_uploader(
                "Subir documentos",
                type=["pdf", "jpg", "jpeg", "png", "xlsx", "xls", "docx"],
                accept_multiple_files=True,
                key="veh_archivos"
            )

            enviado = st.form_submit_button("Guardar",use_container_width=True,width=150,icon="💾")

    col1, col2, col3 = st.columns(3)
    with col2:
        limpiar = st.button("Limpiar Registro", on_click=limpiar_y_recargar,use_container_width=True,width=100,icon="🗑️")
    with col3:
        if st.button("Volver al inicio",icon="⬅️",use_container_width=True,width=100):
            st.session_state.vista = None
            st.rerun()
    # ======================= VALIDACIONES =============================
        if enviado:
            response = (
                supabase
                .table("BitacoraOperaciones")
                .select("*")
                .eq("NUM_SINIESTRO",Siniestro)
                .limit(1)
                .execute()
            )

            if response.data:
                st.toast("El número de expediente ya se encuentra registrado. Use un ID diferente o revise la pestaña “Modificar datos”.", icon="🚨",duration=3)
                return

            errores = []

            if not Siniestro:
                errores.append("El número de siniestro es obligatorio.")

            email_regex = r"^[\w\.-]+@[\w\.-]+\.\w+$"
            if Asegurado_Correo and not re.match(email_regex, Asegurado_Correo):
                errores.append("El correo del asegurado no es válido.")

            if errores:
                st.error("Revisa lo siguiente:\n- " + "\n- ".join(errores))
                return
            
            # Usuario login desde session_state
            Usuario_Login = st.session_state["USUARIO"]
            Liquidador_Nombre = st.session_state["LIQUIDADOR"]

            # Crear carpeta en drive
            carpeta_id = obtener_o_crear_carpeta(f"SINIESTRO_{Siniestro}", drive_service)
            carpeta_link = f"https://drive.google.com/drive/folders/{carpeta_id}"

            # Subir archivos
            links_archivos = []
            if archivos:
                for archivo in archivos:
                    archivo_id = subir_archivo_drive(
                        archivo.name,
                        archivo.read(),
                        archivo.type,
                        carpeta_id,
                        drive_service
                    )
                    links_archivos.append(f"https://drive.google.com/file/d/{archivo_id}/view")

            supabase.table("BitacoraOperaciones").insert({
                "NUM_SINIESTRO": Siniestro,
                "CORRELATIVO": Correlativo,
                "FECHA_SINIESTRO": FechaSiniestro.strftime("%Y-%m-%d"),
                "LUGAR_SINIESTRO": Lugar,
                "MEDIO": Medio,
                "COBERTURA": Cobertura,
                "MARCA": Marca,
                "SUBMARCA": Submarca,
                "VERSION": Version,
                "MODELO": AñoModelo,
                "NO_SERIE": Serie,
                "MOTOR": Motor,
                "PATENTE": Patente,
                "FECHA_CREACION": datetime.now(ZoneInfo("America/Mexico_City")).strftime("%Y-%m-%d"),
                "FECHA_ESTATUS_BITACORA": datetime.now(ZoneInfo("America/Mexico_City")).strftime("%Y-%m-%d %H:%M:%S"),
                "ESTATUS": "ALTA SINIESTRO",
                "NOMBRE_ASEGURADO": Asegurado_Nombre,
                "RUT_ASEGURADO": Asegurado_Rut,
                "TIPO_DE_PERSONA_ASEGURADO": Asegurado_Tipo,
                "TEL_ASEGURADO": Asegurado_Telefono,
                "CORREO_ASEGURADO": Asegurado_Correo,
                "DIRECCION_ASEGURADO": Asegurado_Direccion,
                "NOMBRE_PROPIETARIO": Propietario_Nombre,
                "RUT_PROPIETARIO": Propietario_Rut,
                "TIPO_DE_PERSONA_PROPIETARIO": Propietario_Tipo,
                "TEL_PROPIETARIO": Propietario_Telefono,
                "CORREO_PROPIETARIO": Propietario_Correo,
                "DIRECCION_PROPIETARIO": Propietario_Direccion,
                "LIQUIDADOR": Liquidador_Nombre,
                "CORREO_LIQUIDADOR": Usuario_Login,
                "DRIVE": carpeta_link
            }).execute()

            st.toast("Guardando cambios...", icon="⏳",duration=1)
            time.sleep(1)
            st.toast("Siniestro registrado correctamente", icon="✅")
           #st.success("Siniestro registrado correctamente", icon="✅",width=30)

def vista_buscar_siniestro():

    st.subheader("🔎 Buscar siniestro")

    siniestro = st.text_input("ESCRIBE NÚMERO DE SINIESTRO:")

    if st.button("Buscar", icon="🔎", use_container_width=True):
        
        if not siniestro:
            st.warning("Ingresa un número de siniestro.")
            return
        try:
            response = (
                supabase
                .table("BitacoraOperaciones")
                .select("*")
                .eq("NUM_SINIESTRO", siniestro)
                .execute()
            )
        except Exception as e:
            st.error("Error al consultar el siniestro.")
            st.write(e)
            st.stop()

        if not response.data:
            st.error("Siniestro no encontrado.",icon="❌")
            return

        resultado = pd.DataFrame(response.data)

        if resultado.empty:
            st.error("❌ Siniestro no encontrado.")
            return
        resultado.rename(columns={
            "NUM_SINIESTRO":"# DE SINIESTRO",
            "CORRELATIVO":"CORRELATIVO",
            "FECHA_SINIESTRO":"FECHA SINIESTRO",
            "LUGAR_SINIESTRO":"LUGAR SINIESTRO",
            "MEDIO":"MEDIO ASIGNACIÓN",
            "COBERTURA":"COBERTURA",
            "MARCA":"MARCA",
            "SUBMARCA":"SUBMARCA",
            "VERSION":"VERSIÓN",
            "MODELO":"AÑO/MODELO",
            "NO_SERIE":"NO. SERIE",
            "MOTOR":"MOTOR",
            "PATENTE":"PATENTE",
            "FECHA_CREACION":"FECHA CREACIÓN",
            "FECHA_ESTATUS_BITACORA":"FECHA ESTATUS BITÁCORA",
            "ESTATUS":"ESTATUS",
            "NOMBRE_ASEGURADO":"NOMBRE ASEGURADO",
            "RUT_ASEGURADO":"RUT ASEGURADO",
            "TIPO_DE_PERSONA_ASEGURADO":"TIPO DE PERSONA ASEGURADO",
            "TEL_ASEGURADO":"TEL. ASEGURADO",
            "CORREO_ASEGURADO":"CORREO ASEGURADO",
            "DIRECCION_ASEGURADO":"DIRECCIÓN ASEGURADO",
            "NOMBRE_PROPIETARIO":"NOMBRE PROPIETARIO",
            "RUT_PROPIETARIO":"RUT PROPIETARIO",
            "TIPO_DE_PERSONA_PROPIETARIO":"TIPO DE PERSONA PROPIETARIO",
            "TEL_PROPIETARIO":"TEL. PROPIETARIO",
            "CORREO_PROPIETARIO":"CORREO PROPIETARIO",
            "DIRECCION_PROPIETARIO":"DIRECCIÓN PROPIETARIO",
            "LIQUIDADOR":"LIQUIDADOR",
            "CORREO_LIQUIDADOR":"CORREO LIQUIDADOR",
            "DRIVE":"DRIVE",
            "COMENTARIO":"COMENTARIO"
        },inplace=True)

        
        resultado["FECHA ESTATUS BITÁCORA"] = pd.to_datetime(resultado["FECHA ESTATUS BITÁCORA"], errors="coerce")
        resultado = resultado.sort_values(by=["FECHA ESTATUS BITÁCORA"],ascending=[True])
        Ultimo_estatus = resultado.iloc[-1]["ESTATUS"]
        st.success(f"Último estatus registrado: {Ultimo_estatus}")
        st.info("Registro de operaciones completo:")
        st.dataframe(resultado, use_container_width=True, hide_index=True)

        # ============================
        #   LINK A DRIVE
        # ============================
        if "DRIVE" in resultado.columns:
            drive_link = resultado.iloc[0]["DRIVE"]

            if isinstance(drive_link, str) and drive_link.startswith("http"):
                st.markdown(
                    f"<a href='{drive_link}' target='_blank' style='font-size:18px;'>📁 Abrir carpeta en Drive</a>",
                    unsafe_allow_html=True
                )
            else:
                st.info("Este siniestro no tiene un link de Drive registrado.")

        else:
            st.info("La columna 'DRIVE' no existe en el registro.")

    if st.button("Volver al inicio", icon="⬅️",use_container_width=True):
        st.session_state.vista = None
        st.rerun()

def vista_descargas():
    st.subheader("📥 Descargas")
    response = (
        supabase
        .table("BitacoraOperaciones")
        .select("*")
        .execute()
    )
    resultado = pd.DataFrame(response.data)
    resultado.rename(columns={
        "NUM_SINIESTRO":"# DE SINIESTRO",
        "CORRELATIVO":"CORRELATIVO",
        "FECHA_SINIESTRO":"FECHA SINIESTRO",
        "LUGAR_SINIESTRO":"LUGAR SINIESTRO",
        "MEDIO":"MEDIO ASIGNACIÓN",
        "COBERTURA":"COBERTURA",
        "MARCA":"MARCA",
        "SUBMARCA":"SUBMARCA",
        "VERSION":"VERSIÓN",
        "MODELO":"AÑO/MODELO",
        "NO_SERIE":"NO. SERIE",
        "MOTOR":"MOTOR",
        "PATENTE":"PATENTE",
        "FECHA_CREACION":"FECHA CREACIÓN",
        "FECHA_ESTATUS_BITACORA":"FECHA ESTATUS BITÁCORA",
        "ESTATUS":"ESTATUS",
        "NOMBRE_ASEGURADO":"NOMBRE ASEGURADO",
        "RUT_ASEGURADO":"RUT ASEGURADO",
        "TIPO_DE_PERSONA_ASEGURADO":"TIPO DE PERSONA ASEGURADO",
        "TEL_ASEGURADO":"TEL. ASEGURADO",
        "CORREO_ASEGURADO":"CORREO ASEGURADO",
        "DIRECCION_ASEGURADO":"DIRECCIÓN ASEGURADO",
        "NOMBRE_PROPIETARIO":"NOMBRE PROPIETARIO",
        "RUT_PROPIETARIO":"RUT PROPIETARIO",
        "TIPO_DE_PERSONA_PROPIETARIO":"TIPO DE PERSONA PROPIETARIO",
        "TEL_PROPIETARIO":"TEL. PROPIETARIO",
        "CORREO_PROPIETARIO":"CORREO PROPIETARIO",
        "DIRECCION_PROPIETARIO":"DIRECCIÓN PROPIETARIO",
        "LIQUIDADOR":"LIQUIDADOR",
        "CORREO_LIQUIDADOR":"CORREO LIQUIDADOR",
        "DRIVE":"DRIVE",
        "COMENTARIO":"COMENTARIO"
    },inplace=True)

    st.write("Selecciona el tipo de bitácora a descargar.")

    opcion = st.selectbox(
        "Tipo de descarga",
        [
            "Selecciona una opción",
            "Bitácora de operación",
            "Bitácora de último estatus"
        ],
        key="tipo_descarga"
    )
    # --- BITÁCORA DE OPERACIÓN ---
    if opcion == "Bitácora de operación":

        resultado["FECHA ESTATUS BITÁCORA"] = pd.to_datetime(resultado["FECHA ESTATUS BITÁCORA"],errors="coerce")
        resultado["FECHA SINIESTRO"] = pd.to_datetime(resultado["FECHA SINIESTRO"], errors="coerce")
        resultado["FECHA CREACIÓN"] = pd.to_datetime(resultado["FECHA CREACIÓN"], errors = "coerce")
        
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            resultado.to_excel(writer, index=False, sheet_name="LOG")
            ws = writer.book["LOG"]
            header_fill = PatternFill("solid", fgColor="1F4E78")  # azul
            header_font = Font(bold=True, color="FFFFFF")
            header_align = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            for col_idx, cell in enumerate(ws[1], start=1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                ws.column_dimensions[get_column_letter(col_idx)].width = 22
            ws.row_dimensions[1].height = 35

        st.download_button(
            label="Descargar bitácora de operación",
            icon="⬇️",
            use_container_width=True,
            disabled=False,
            data=buffer.getvalue(),
            file_name="Bitacora_Operación_SURA.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # --- BITÁCORA DE ÚLTIMO ESTATUS ---
    elif opcion == "Bitácora de último estatus":
        resultado["FECHA ESTATUS BITÁCORA"] = pd.to_datetime(resultado["FECHA ESTATUS BITÁCORA"],errors="coerce")
        resultado["FECHA SINIESTRO"] = pd.to_datetime(resultado["FECHA SINIESTRO"], errors="coerce")
        resultado["FECHA CREACIÓN"] = pd.to_datetime(resultado["FECHA CREACIÓN"], errors = "coerce")

        df_sorted = resultado.sort_values(
            by=["# DE SINIESTRO", "FECHA ESTATUS BITÁCORA"],
            ascending=[True, True]
        )

        df_ultimos = df_sorted.groupby("# DE SINIESTRO").tail(1)

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df_ultimos.to_excel(writer, index=False, sheet_name="LOG")
            ws = writer.book["LOG"]
            header_fill = PatternFill("solid", fgColor="1F4E78")  # azul
            header_font = Font(bold=True, color="FFFFFF")
            header_align = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            for col_idx, cell in enumerate(ws[1], start=1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                ws.column_dimensions[get_column_letter(col_idx)].width = 22
            ws.row_dimensions[1].height = 35

        st.download_button(
            label="Descargar bitácora de último estatus",
            icon="⬇️",
            use_container_width=True,
            disabled=False,
            data=buffer.getvalue(),
            file_name="Bitacora_UltimoEstatus_SURA.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    if st.button("Volver al inicio",icon="⬅️",use_container_width=True,width=100):
        st.session_state.vista = None
        st.rerun()

def vista_registro_usuario():
    st.header("Registro de nuevo usuario")

    with st.form("form_usuario",clear_on_submit=True):
        usuario = st.text_input("NOMBRE COMPLETO:", key="nom_usuario")
        correo = st.text_input("CORREO:", key="correo")
        password = st.text_input("CONTRASEÑA", key="password",type="password")
        rol = st.selectbox("ROL",["Seleccionar rol","ADMINISTRADOR","LIQUIDADOR"],key="rol")

        enviado = st.form_submit_button("Guardar datos",use_container_width=True,width=150,icon="💾")

        if enviado:
            errores = []

            if not usuario:
                errores.append("Falta ingresar nombre del usuario")
            if not correo:
                errores.append("Falta ingresar correo del usuario")
            if not password:
                errores.append("Falta ingresar contraseña del usuario")
            if rol == "Seleccionar rol":
                errores.append("Seleccionar un rol para el usuario")

            email_regex = r"^[\w\.-]+@[\w\.-]+\.\w+$"
            if correo and not re.match(email_regex, correo):
                errores.append("El correo del usuario no es válido.")

            if errores:
                st.error("Revisa lo siguiente:\n- " + "\n- ".join(errores))
                return
            
            response = (
                supabase
                .table("Login")
                .select("USUARIO","PASSWORD","ROL","LIQUIDADOR")
                .eq("USUARIO",correo)
                .limit(1)
                .execute()
            )

            if response.data:
                st.error("La dirección de correo ingresada ya se encuentra asociada a un perfil de liquidador. Intenta nuevamente con una cuenta distinta.", icon="🚨")
                return

            supabase.table("Login").insert({
                "USUARIO": correo,
                "PASSWORD": bcrypt.hashpw(str(password).encode("utf-8"),bcrypt.gensalt()).decode("utf-8"),
                "ROL": rol,
                "LIQUIDADOR": usuario.upper()
            }).execute()

            st.success("Usuario registrado correctamente")
            CLAVE_APP = 'ckkazcijqkwikscd' #Contraseña de aplicación, utilizada para acceder al correo
            REMITENTE = 'jbarron@cibergestion.com' 
            DESTINATARIO = correo
            ASUNTO = 'SURA PT - CREACIÓN DE USUARIO'
            MENSAJE = f"""Estimado <strong>{usuario}</strong>,
            Tu usuario se ha creado correctamente. A continuación, se detallan tus datos de acceso:

            <strong>USUARIO: </strong>{correo}
            <strong>CONTRASEÑA: </strong>{password}
            <strong>ROL:</strong>{rol}

            Puedes acceder a la plataforma haciendo clic en el siguiente enlace: <a href="https://sura-pt-cibergestion.streamlit.app/">Ingresar aquí</a>
            
            Saludos.
            """
            yag = yagmail.SMTP(REMITENTE, CLAVE_APP)

            #Se envía el correo a los destinatarios con la estructura establecida en el item contents

            yag.send(
                to=DESTINATARIO, 
                subject=ASUNTO, 
                contents=[MENSAJE]
                )
        
    if st.button("Volver al inicio",icon="⬅️",use_container_width=True,width=100):
        st.session_state.vista = None
        st.rerun()
# =======================================================
#               DASHBOARD GENERAL Y PARTICULAR
# =======================================================
def kpi_card_2(titulo, valor, color_bg, color_font):
    st.markdown(
        f"""
        <div style="
            background-color:{color_bg};
            padding:16px;
            border-radius:12px;
            text-align:center;
            color:{color_font};
            height:140px;
            display:flex;
            flex-direction:column;
            justify-content:left;
        ">
            <span style="font-size:20px; font-weight:500; color:{color_font};"><strong>{titulo}</strong></span>
            <h2 style="margin:0;">{valor}</h2>
        </div>
        """,
        unsafe_allow_html=True
    )

def kpi_card(backcolor, border, text1, text2, valor):
    st.markdown(
        f"""
        <div style="
            background-color:{backcolor};
            padding:5px;
            border-left:7px solid {border};
            border-radius:6px;
            font-size:20px;
            text-align:center;
            color:"black";
        ">
            <p style="font-size: 20px;"><strong>{text1}</strong></p>
            <p style="font-size: 25px; color: rgb(172, 0, 0);"<strong><font color="red">{valor}</strong></font></p>
            <p style="font-size: 20px;"><strong>{text2}</strong></p>
        </div>
        """,
        unsafe_allow_html=True
        )

def dash_general():
    Liquidador_Nombre = st.session_state["LIQUIDADOR"]
    st.header(f"¡HOLA, {Liquidador_Nombre}!")

    response = (
    supabase
    .table("BitacoraOperaciones")
    .select("*")
    .execute()
    )
    df_dash = pd.DataFrame(response.data)
    df_dash["FECHA_ESTATUS_BITACORA"] = pd.to_datetime(df_dash["FECHA_ESTATUS_BITACORA"],errors="coerce")
    df_dash["FECHA_CREACION"] = pd.to_datetime(df_dash["FECHA_CREACION"], errors="coerce")
    df_dash = (df_dash.sort_values(by=["NUM_SINIESTRO", "FECHA_ESTATUS_BITACORA"],ascending=[True, True]).groupby("NUM_SINIESTRO").tail(1))

    estatus_cierre = [
    "PAGO LIBERADO",
    "CIERRE POR DESISTIMIENTO",
    "CIERRE POR RECHAZO",
    "SOLICITUD DE PAGO GENERADA"
    ]

    df_cerrados = df_dash[df_dash["ESTATUS"].isin(estatus_cierre)]
    df_cerrados["DIAS_HABILES"] = np.busday_count(df_cerrados["FECHA_CREACION"].values.astype("datetime64[D]"), 
                                                  df_cerrados["FECHA_ESTATUS_BITACORA"].values.astype("datetime64[D]"))

    #df_cerrados = df_dash[(df_dash["ESTATUS"] == "PAGO LIBERADO") | (df_dash["ESTATUS"] == "SOLICITUD DE PAGO GENERADA")]
    
    total_siniestros = df_dash.shape[0]
    total_cerrados = df_cerrados.shape[0]
    Per_cerrados = str(int((total_cerrados/total_siniestros)*100)) + " %"
    promedio_dias_cierre = round(df_cerrados["DIAS_HABILES"].mean(), 1)


    st.subheader("MÉTRICAS GENERALES",divider="blue")
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        #kpi_card("SINIESTROS RECIBIDOS",total_siniestros,"#FFF4EA","#FBA21B")
        kpi_card("#FFF4EA","#FBA21B","SINIESTROS","RECIBIDOS",total_siniestros)
    with col2:
        #kpi_card("SINIESTROS CERRADOS",total_cerrados,"#F0FFEA","#047A1B")
        kpi_card("#F0FFEA","#047A1B","SINIESTROS","CERRADOS",total_cerrados)
    with col3:
        #kpi_card("% CERRADOS", Per_cerrados, "#E6F1FD","#6DA1AF")
        kpi_card("#E6F1FD","#6DA1AF","%","CERRADOS",Per_cerrados)
    with col4:
        #kpi_card("DÍAS PROMEDIO CIERRE", promedio_dias_cierre, "#E6F1FD","#6DA1AF")
        kpi_card("#FFF5F7","#A82A50","DÍAS PROMEDIO","CIERRE",promedio_dias_cierre)

    #Agregamos gráficas generales
    st.divider()
    col1, col2 = st.columns(2)

    count_estatus = df_dash.groupby("ESTATUS").size().reset_index(name="TOTAL")
    count_liquidador = df_dash.groupby("LIQUIDADOR").size().reset_index(name="TOTAL")

    with col1:
        st.markdown("### TOTAL DE SINIESTROS POR ESTATUS")

        chart_estatus = alt.Chart(count_estatus).mark_bar().encode(
            x=alt.X("TOTAL:Q", title="TOTAL"),
            y=alt.Y("ESTATUS:N", sort="-x", title=""),
            color=alt.value("#0fb3ff86"),
            tooltip=["ESTATUS", "TOTAL"]
        ).properties(height=400)

        st.altair_chart(chart_estatus, use_container_width=True)
    with col2:
        st.markdown("### TOTAL DE SINIESTROS POR LIQUIDADOR")

        chart_liquidador = alt.Chart(count_liquidador).mark_bar().encode(
            x=alt.X("TOTAL:Q", title="TOTAL"),
            y=alt.Y("LIQUIDADOR:N", sort="-x", title=""),
            color=alt.value("#992cff86"),
            tooltip=["LIQUIDADOR", "TOTAL"]
        ).properties(height=400)

        st.altair_chart(chart_liquidador, use_container_width=True)
    
def dash_liquidador():
    st.divider()
    Liquidador = st.session_state["LIQUIDADOR"]
    response = (
    supabase
    .table("BitacoraOperaciones")
    .select("*")
    .eq("LIQUIDADOR", Liquidador)
    .execute()
    )
    if not response.data:
        st.subheader("MÉTRICAS PARTICULARES",divider="blue")

        st.markdown(
        """
        <div style="
            background-color:#f3f4f6;
            padding:12px;
            border-left:5px solid #3b82f6;
            border-radius:6px;
            font-size:14px;
        ">
            📄 <strong>Último estatus de tus siniestros asignados.</strong><br>
            Sin registros asignados.
        </div>
        """,
        unsafe_allow_html=True
        )
    else:
        df_dash = pd.DataFrame(response.data)
        df_dash["FECHA_ESTATUS_BITACORA"] = pd.to_datetime(df_dash["FECHA_ESTATUS_BITACORA"],errors="coerce")
        df_dash = (df_dash.sort_values(by=["NUM_SINIESTRO", "FECHA_ESTATUS_BITACORA"],ascending=[True, True]).groupby("NUM_SINIESTRO").tail(1))
        df_dash.rename(columns={
                "NUM_SINIESTRO":"# DE SINIESTRO",
                "CORRELATIVO":"CORRELATIVO",
                "FECHA_SINIESTRO":"FECHA SINIESTRO",
                "LUGAR_SINIESTRO":"LUGAR SINIESTRO",
                "MEDIO":"MEDIO ASIGNACIÓN",
                "COBERTURA":"COBERTURA",
                "MARCA":"MARCA",
                "SUBMARCA":"SUBMARCA",
                "VERSION":"VERSIÓN",
                "MODELO":"AÑO/MODELO",
                "NO_SERIE":"NO. SERIE",
                "MOTOR":"MOTOR",
                "PATENTE":"PATENTE",
                "FECHA_CREACION":"FECHA CREACIÓN",
                "FECHA_ESTATUS_BITACORA":"FECHA ESTATUS BITÁCORA",
                "ESTATUS":"ESTATUS",
                "NOMBRE_ASEGURADO":"NOMBRE ASEGURADO",
                "RUT_ASEGURADO":"RUT ASEGURADO",
                "TIPO_DE_PERSONA_ASEGURADO":"TIPO DE PERSONA ASEGURADO",
                "TEL_ASEGURADO":"TEL. ASEGURADO",
                "CORREO_ASEGURADO":"CORREO ASEGURADO",
                "DIRECCION_ASEGURADO":"DIRECCIÓN ASEGURADO",
                "NOMBRE_PROPIETARIO":"NOMBRE PROPIETARIO",
                "RUT_PROPIETARIO":"RUT PROPIETARIO",
                "TIPO_DE_PERSONA_PROPIETARIO":"TIPO DE PERSONA PROPIETARIO",
                "TEL_PROPIETARIO":"TEL. PROPIETARIO",
                "CORREO_PROPIETARIO":"CORREO PROPIETARIO",
                "DIRECCION_PROPIETARIO":"DIRECCIÓN PROPIETARIO",
                "LIQUIDADOR":"LIQUIDADOR",
                "CORREO_LIQUIDADOR":"CORREO LIQUIDADOR",
                "DRIVE":"DRIVE",
                "COMENTARIO":"COMENTARIO"
            },inplace=True)

        st.subheader("MÉTRICAS PARTICULARES",divider="blue")

        st.markdown(
        """
        <div style="
            background-color:#f3f4f6;
            padding:12px;
            border-left:5px solid #3b82f6;
            border-radius:6px;
            font-size:14px;
        ">
            📄 <strong>Último estatus de tus siniestros asignados.</strong><br>
            Puedes descargar la tabla haciendo clic en el ícono de descarga en la esquina superior derecha.
        </div>
        """,
        unsafe_allow_html=True
        )
        
        st.divider()
        st.dataframe(data=df_dash,hide_index=True,)



# =======================================================
#               VISTA LIQUIDADOR
# =======================================================

def vista_liquidador():

    Liquidador_Nombre = st.session_state["LIQUIDADOR"]
    st.title(f"¡HOLA, {Liquidador_Nombre}!")

    # Inicializar la variable si no existe
    if "vista" not in st.session_state:
        st.session_state.vista = None
    # ----------------------------------------------------
    #  MENÚ LATERAL
    # ----------------------------------------------------
    with st.sidebar.expander("GESTIÓN DE SINIESTRO", expanded=False):
        if st.button("REGISTRAR", use_container_width=True, icon="📄"):
            st.session_state.vista = "REGISTRAR"

        if st.button("ACTUALIZAR", use_container_width=True, icon="🔄️"):
            st.session_state.vista = "ACTUALIZAR"

        if st.button("SUBIR ARCHIVOS", use_container_width=True, icon="⬆️"):
            st.session_state.vista = "CARGA"
    with st.sidebar:
        if st.button("BUSCAR / CONSULTAR", use_container_width=True, icon="🔎"):
            st.session_state.vista = "BUSCAR"

        if st.button("Cerrar sesión", use_container_width=True,icon="❌"):
            st.session_state.clear()
            st.rerun()
    # =====================================================================================
    #                                REGISTRAR SINIESTRO
    # =====================================================================================
    if st.session_state.vista == "REGISTRAR":
        registro_siniestro()

    elif st.session_state.vista == "ACTUALIZAR":
        vista_modificar_siniestro()
    
    elif st.session_state.vista == "CARGA":
        panel_subir_documentos()

    elif st.session_state.vista == "BUSCAR":
        vista_buscar_siniestro()

    elif st.session_state.vista == None:
        dash_general()
        dash_liquidador()
    

# =======================================================
#                VISTA ADMINISTRADOR
# =======================================================
def vista_admin():
    Liquidador_Nombre = st.session_state["LIQUIDADOR"]
    #st.title(f"¡HOLA, {Liquidador_Nombre}!")
    image_url = "https://www.segurossura.com.co/boletincovid191/Recomendaciones_COVID-19/images/logo_sura.png"
    st.logo(image=image_url,size="large")

    # Inicializar la variable si no existe
    if "vista" not in st.session_state:
        st.session_state.vista = None
    # ----------------------------------------------------
    #  MENÚ LATERAL
    # ----------------------------------------------------
    with st.sidebar.expander("ADMINISTRACIÓN", expanded=False):
        if st.button("DESCARGAS", use_container_width=True, icon="📥"):
            st.session_state.vista = "DESCARGA"

        if st.button("USUARIOS", use_container_width=True, icon="👥"):
            st.session_state.vista = "USUARIOS"

    with st.sidebar.expander("GESTIÓN DE SINIESTRO", expanded=False):
        if st.button("REGISTRAR", use_container_width=True, icon="📄"):
            st.session_state.vista = "REGISTRAR"

        if st.button("ACTUALIZAR", use_container_width=True, icon="🔄️"):
            st.session_state.vista = "ACTUALIZAR"

        if st.button("SUBIR ARCHIVOS", use_container_width=True, icon="⬆️"):
            st.session_state.vista = "CARGA"

    with st.sidebar:
        if st.button("BUSCAR / CONSULTAR", use_container_width=True, icon="🔎"):
            st.session_state.vista = "BUSCAR"
        #st.markdown("<div style='height:20vh'></div>", unsafe_allow_html=True)
        if st.button("Cerrar sesión", use_container_width=True,icon="❌"):
            st.session_state.clear()
            st.rerun()
    # =====================================================================================
    #                                REGISTRAR SINIESTRO
    # =====================================================================================
    if st.session_state.vista == "REGISTRAR":
        registro_siniestro()
    elif st.session_state.vista == "ACTUALIZAR":
        vista_modificar_siniestro()
    elif st.session_state.vista == "CARGA":
        panel_subir_documentos()
    elif st.session_state.vista == None:
        dash_general()
        dash_liquidador()

    # =====================================================================================
    #                                BUSCAR / ACTUALIZAR
    # =====================================================================================

    elif st.session_state.vista == "BUSCAR":
        vista_buscar_siniestro()
    elif st.session_state.vista == "DESCARGA":
        vista_descargas()
    elif st.session_state.vista == "USUARIOS":
        vista_registro_usuario()

# =======================================================
#                 CONTROL DE SESIÓN
# =======================================================
if "auth" not in st.session_state:
    st.session_state["auth"] = False

if not st.session_state["auth"]:
    login()
    st.stop()

# =======================================================
#                 INTERFAZ PRINCIPAL
# =======================================================

with st.sidebar:
    st.markdown("""
        <div style="text-align:center; padding-top:10px;">
            <h1 style="margin-bottom:0; color: #00AEC7;">SURA PÉRDIDAS TOTALES</h1>
        </div>
    """, unsafe_allow_html=True)

    st.subheader("📋 MENÚ")

if st.session_state["ROL"] == "ADMINISTRADOR":
    vista_admin()
else:
    vista_liquidador()
